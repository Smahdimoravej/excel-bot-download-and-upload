from typing import Text
from openpyxl import *
from openpyxl.drawing.image import Image
import random
import datetime
from telegram import *
from telegram import update
from telegram.ext import *
from telegram.ext import callbackcontext
token = "1895042623:AAFpxW23oTtvIuIF70B6LoqNnW9vwbsT6Kw"
bot = Bot(token)
updater = Updater(token, use_context=True)
dispatcher: Dispatcher = updater.dispatcher
d = datetime.datetime.now()
wb = Workbook()
keyboard_1 = [
            [InlineKeyboardButton(
             "چاه",
             callback_data="chah")]
             ]
keyboard_2 = [
            [InlineKeyboardButton(
             "جعفریه",
             callback_data="جعفریه")]
             ]
keyboard_3 = [
            [InlineKeyboardButton(
             "چاه وسفونجرد",
             callback_data="چاه وسفونجرد جعفریه36010101.xlsx")]
            ]
reply_mark_1 = InlineKeyboardMarkup(keyboard_1)
reply_mark_2 = InlineKeyboardMarkup(keyboard_2)
reply_mark_3 = InlineKeyboardMarkup(keyboard_3)


def messagetoUs(update: Update, context: CallbackContext):
    print(update)
    try:
        bot.send_message(
            chat_id=update.message.chat_id,
            text='''
                این ربات جهت راحتی پیمانکار طراحی شده است.
                 ''',
            reply_markup=reply_mark_1,
            )
    except Exception as e:
        print("error")


def query_btns(update: Update, context: CallbackContext):
    query: CallbackQuery = update.callback_query
    if query.data == "chah":
        bot.send_message(
            chat_id=update.effective_message.chat_id,
            text="222",
            reply_markup=reply_mark_2
            )
    elif query.data == "جعفریه":
        bot.send_message(
            chat_id=update.effective_message.chat_id,
            text="2222",
            reply_markup=reply_mark_3
            )
    elif query.data == "چاه وسفونجرد جعفریه36010101.xlsx":
        # bot.send_message(
        #     chat_id=update.effective_message.chat_id,
        #     text="چاه وسفونجرد جعفریه36010101.xlsx",           
        #     # reply_markup=reply_mark_3
        #     )
        # input = update.message.text
        # context.bot.send_message(chat_id=update.effective_chat.id, text=msg)
        wb = load_workbook(query.data)
        ws = wb.active
        # ws['a5'] = update.
        wb.save(filename="چاه وسفونجرد جعفریه36010101.xlsx")


dispatcher.add_handler(MessageHandler(Filters.all, messagetoUs))
dispatcher.add_handler(CallbackQueryHandler(query_btns))
updater.start_polling()
